#!/usr/bin/env python3
"""
Module for tasks that do post-run processing of output files.
"""
import logging
import subprocess
from os import remove
from datetime import datetime
from typing import Generator, List, NamedTuple, Tuple

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from framework.tasks.data_classes import RecordContext

OLINK_FIRST_COLUMN = ["npx data", "panel", "assay", "uniprot id", "olinkid"]

MANIFEST_HEADERS = {
    "manifest id:": "manifest_id",
    "protocol id:": "protocol_id",
    "request:": "request",
    "assay priority:": "assay_priority",
    "assay type:": "assay_type",
    "batch number:": "batch_number"
}

SHIPPING_DETAILS = {
    "shipping condition:": "shipping_condition",
    "date shipped:": "date_shipped",
    "# of samples shipped:": "number_shipped",
}

SENDER_ADDRESS = {
    "name:": "sender_name",
    "address:": "sender_address",
    "email:": "sender_email",
}
RECEIVER_ADDRESS = {
    "name:": "receiver_name",
    "address:": "receiver_address",
    "email:": "receiver_email",
}
SHIPPER_INFO = {
    "courier:": "courier",
    "tracking number:": "tracking_number",
    "account number:": "account_number",
}
SAMPLE_DESCRIPTION = {
    "pathology report": "pathology_report",
    "time point": "time_point",
    "specimen type": "specimen_type",
    "specimen format": "specimen_format",
    "collection date": "collection_date",
    "processing date": "processing_date",
    "quantity": "quantity",
    "volume": "volume",
    "units": "units",
    "sample source": "sample_source",
    "comments": "comments",
}

FIELD_NAME_LIST = [
    (MANIFEST_HEADERS, 1),
    (SHIPPER_INFO, 1),
    (SHIPPING_DETAILS, 4),
    (SENDER_ADDRESS, 2),
    (RECEIVER_ADDRESS, 5),
]


class SampleInfo(NamedTuple):
    """
    Basic store for sample related info.

    Attributes:
        plate_gen {Generator} -- plate info
        qc_gen {Generator} -- qc info
        sample_ids {List[str]} -- id list

    Arguments:
        NamedTuple {[type]} -- [description]
    """

    plate_gen: Generator
    qc_gen: Generator
    sample_ids: List[str]


class AssayColumn(NamedTuple):
    """
    Structure for holding all the generators needed to get assay information
        from olink files.
    """

    assay_info_gen: Generator
    assay_metrics_gen: Generator
    data_col_gen: Generator


def mk_error(
    explanation: str,
    affected_paths: List[str],
    raw_or_parse: str = "PARSE",
    severity: str = "WARNING",
) -> dict:
    """
    Handy function for creating a validation error.

    Arguments:
        explanation {str} -- Long form explanation of the error.
        affected_paths {List[str]} -- List of the json paths in the mongo record affected by the
            error.

    Keyword Arguments:
        raw_or_parse {str} -- Whether the issue is with the file object or the content.
            (default: {"PARSE"})
        severity {str} -- How severe the error is. (default: {"WARNING"})

    Returns:
        dict -- Validation error.
    """
    return {
        "explanation": explanation,
        "affected_paths": affected_paths,
        "raw_or_parse": raw_or_parse,
        "severity": severity,
    }


def add_file_extension(path: str, extension: str) -> str:
    """
    Adds a specified file extension to a given file.

    Arguments:
        path {str} -- Path of the file.
        extension {str} -- Extension to be added.

    Returns:
        str -- Path of the new file.
    """
    newpath = "%s.%s" % (path, extension)
    args = ["mv", path, newpath]
    subprocess.run(args)
    return newpath


def lazy_remove(path: str) -> None:
    """
    Function to remove a file and simply continue if that file has already been deleted.

    Arguments:
        path {str} -- File path.

    Returns:
        None -- [description]
    """
    try:
        remove(path)
    except FileNotFoundError:
        pass


def diff_fields(actual: List[str], expected: List[str]) -> List[str]:
    """
    Compares two lists and returns the items not present in the second list.

    Arguments:
        actual {List[str]} -- Items found in the file.
        expected {List[str]} -- Items expected.

    Returns:
        List[str] -- Set difference of lists.
    """
    return [absent for absent in actual if absent not in expected]


def bad_xlsx(record_errs: List[dict], path: str, err: InvalidFileException) -> None:
    """
    Reporting function for when the xlsx file can't even be opened.

    Arguments:
        record {List[dict]} -- Mongo record.
        path {str} -- File path.
        err {InvalidFileException} -- Error message.

    Returns:
        None -- [description]
    """
    log = (
        "Error loading file %s. This file is not a valid xlsx file. Error message: %s"
        % (path, str(err))
    )
    logging.warning({"message": log, "category": "WARNING-CELERY-FAIR-IMPORT"})
    record_errs.append(
        mk_error(
            "The file was not recognized as a valid xlsx sheet and could not be loaded.",
            [],
            raw_or_parse="RAW",
            severity="CRITICAL",
        )
    )


def validate_qc_info(generators: SampleInfo, errors: List[dict]) -> List[dict]:
    """
    Loops over the qc/plateid info and returns it in a formatted list.

    Arguments:
        generators {SampleInfo} -- Tuple for holding the generators that loop over the cells
            where the QC info is stored.
        errors {List[dict]} -- List of errors.

    Returns:
        List[dict] -- List of SAMPLE_SCHEMA structured dicts.
    """
    sample_list = []
    try:
        # Generator should have a single column.
        plate_col = next(generators.plate_gen)
        qc_col = next(generators.qc_gen)
        for plate_cell, qc_cell, sample_id in zip(
            plate_col, qc_col, generators.sample_ids
        ):
            sample_list.append(
                {
                    "sample_id": sample_id,
                    "qc_status": qc_cell.value,
                    "plate_id": plate_cell.value,
                }
            )
    except StopIteration:
        errors.append(
            mk_error(
                "Unable to find the plate or qc column",
                ["samples.qc_status", "samples.plate_id"],
                raw_or_parse="PARSE",
                severity="WARNING",
            )
        )

    return sample_list


def extract_assay_data(
    assay_obj: AssayColumn, sample_ids: List[str], errors: List[dict]
) -> List[dict]:
    """
    Takes a generator of assay info cells and data containing cells, and returns the
    data in a formatted manner.

    Arguments:
        assay_obj {AssayColumn} -- namedtuple that holds all the generators for data processing.
        sample_ids {List[str]} -- List of sample_ids.
        errors {List[dict]} -- A list of parsing errors that have been found.

    Raises:
        InvalidFileException -- Indicates to the calling function that this file is invalid.

    Returns:
        List[dict] -- List of olink assay structured dicts.
    """
    assay_list = []
    for assay_def, data, metrics in zip(
        assay_obj.assay_info_gen, assay_obj.data_col_gen, assay_obj.assay_metrics_gen
    ):
        # Create the assay definition.
        olink_assay = {
            "panel": assay_def[0].value,
            "assay": assay_def[1].value,
            "uniprot_id": assay_def[2].value,
            "olink_id": assay_def[3].value,
            "lod": metrics[0].value,
            "missing_data_freq": metrics[1].value,
            "results": [],
        }

        # Sanity check sample number.
        if len(data) != len(sample_ids):
            errors.append(
                mk_error(
                    "There is a mismatch between the number of samples and the"
                    "number of data points for assay %s" % olink_assay["assay"],
                    ["samples"],
                    raw_or_parse="PARSE",
                    severity="CRITICAL",
                )
            )
            return []

        # Create a new entry in the assay list.
        # Note that the check for qc/lod checks for any variation from a black-text no-fill
        # cell. If other colors are used to indicate other things, this implementation will
        # have to be changed.
        for i, data_cell in enumerate(data):
            value = float(data_cell.value) if data_cell.value else None
            olink_assay["results"].append(
                {
                    "sample_id": sample_ids[i],
                    "value": value,
                    "qc_fail": data_cell.font.color is not None,
                    "below_lod": data_cell.fill.bgColor.rgb != "00000000",
                }
            )
        assay_list.append(olink_assay)
    return assay_list


def process_sample_rows(cell_gen: Generator) -> Tuple[int, int, int, int, List[str]]:
    """
    Takes a generator that loops over the cells containing the collected data, returns
    relevant coordinates used to further process the file.

    Arguments:
        cell_gen {Generator} -- Generator object to loop over cells containing sample data.
            This generator should be created on the first column, but then iterated over until
            olink ID has been seen.

    Returns:
        Tuple[int, int, int, int, List[str]] -- Tuple containing in order, row of first sample,
            row of last sample, row of missing data frequency, panel type row, and list of sample
            ids.
    """
    mdf_row = None
    sample_cells = []
    sample_ids = []
    panel_type_cell = None

    for sam_cell in cell_gen:
        value = sam_cell[0].value
        if value == "Missing Data freq.":
            mdf_row = sam_cell[0].row
        # If a value is a sampleID, note it down.
        elif value == "Panel":
            # Quick way to get the panel name
            panel_type_cell = sam_cell[0].row
        elif value and value != "LOD":
            sample_cells.append(sam_cell[0])
            sample_ids.append(value)

    return (
        sample_cells[0].row,
        sample_cells[-1].row,
        mdf_row,
        panel_type_cell,
        sample_ids,
    )


def validate_row(row_details: Tuple[List[str], int, int], wks_record: dict) -> None:
    """
    Validates a row of field names has the expected values.

    Arguments:
        row_details {Tuple[List[str], int, int]} -- Field names, row index, column index
            (both 1 based).
        wks_record {dict} -- Record.

    Returns:
        None -- [description]
    """
    try:
        extracted_sample_headers = [
            str(cell[0].value).lower()
            for cell in wks_record["wks"].iter_cols(
                min_col=row_details[1],
                max_col=row_details[1] + len(row_details[0]) - 1,
                min_row=row_details[2],
                max_row=row_details[2],
            )
        ]

        if not extracted_sample_headers == row_details[0]:
            diff = diff_fields(extracted_sample_headers, row_details[0])
            wks_record["record"]["validation_errors"].append(
                mk_error(
                    "Field names for row did not match expected. Missing values: %s"
                    % ", ".join(diff)
                    if diff
                    else "None",
                    [],
                )
            )
    except TypeError:
        wks_record["record"]["validation_errors"].append(
            mk_error("unexpected bad value for sample row headers", [])
        )


def validate_column(
    column_details: Tuple[List[str], int], wks_record: dict
) -> Tuple[int, int]:
    """
    Compares the values of a column of cell with the expected values for those cells.
    If they match, returns the row value of the first and last cells in the column.

    Arguments:
        expected {Tuple[List[str], int]} -- The values, in order, you expect the cells to have,
            and the column index.
        wks_record {dict} -- Record and worksheet

    Returns:
        Tuple[int, int] -- Row of first cell, row of last cell, 1 based.
    """
    start_row = None
    end_row = None
    values = []

    try:
        expected = [key for key in column_details[0]]
        gen_exp = wks_record["wks"].iter_rows(
            min_col=column_details[1], max_col=column_details[1]
        )
        while not end_row:
            cell = next(gen_exp)[0]
            value = str(cell.value) if cell.value else ""
            values.append(value.lower())

            if value.lower() == expected[0].lower():
                values = values[-1:]
                start_row = cell.row
            elif value.lower() == expected[-1].lower():
                end_row = cell.row

        if not values == expected:
            diff = diff_fields(values, expected)
            wks_record["record"]["validation_errors"].append(
                mk_error(
                    "Values of column did not match expected values. Missing: %s"
                    % ", ".join(diff)
                    if diff
                    else "None",
                    diff,
                )
            )

    except StopIteration:
        wks_record["record"]["validation_errors"].append(
            mk_error(
                "Could not find a column matching the provided description. Affected Column: %s"
                % column_details[1],
                [],
            )
        )

    return start_row, end_row


def cast_cell_value(cell) -> object:
    """
    Function that ensures openpyxl cells are returned as the correct data type.

    Arguments:
        cell {openpyxl.cell} -- Single cell from an xlsx worksheet.

    Returns:
        object -- The value, cast to int, float, or returned as is.
    """
    if cell.value:
        if cell.data_type == "n":
            if isinstance(cell.value, float):
                return float(cell.value)
            if isinstance(cell.value, int):
                return int(cell.value)
        if isinstance(cell.value, datetime):
                return str(cell.value)
        return cell.value
    return None


def extract_cell_values(generator: Generator) -> List[object]:
    """
    Given a single-column generator expression,
    extract all rows from the generator.

    Arguments:
        generator {Generator} -- Single column generator exprsesion for xlsx worksheet.

    Returns:
        List[object] -- List of the values of the cells.
    """
    values = []
    for cell in generator:
        values.append(cast_cell_value(cell[0]))
    return values


def parse_matched_column(
    column_details: Tuple[List[str], int], wks_record: dict
) -> dict:
    """
    Extracts information from side by side name-value columns.

    Arguments:
        column_details {Tuple[List[str], int]} -- List of field names, column index (1 base)
        wks_record {dict} -- Record object.

    Returns:
        dict -- Column information in dict form.
    """
    start_row, end_row = validate_column(column_details, wks_record)
    values = extract_cell_values(
        wks_record["wks"].iter_rows(
            min_col=column_details[1] + 1,
            max_col=column_details[1] + 1,
            min_row=start_row,
            max_row=end_row,
        )
    )
    data = {}
    for field_name, value in zip(
        [value for key, value in column_details[0].items()], values
    ):
        data[field_name] = value

    return data


def find_row_wise(wks_record: dict, header_fields: List[str]) -> Tuple[int, int]:
    """
    Looks for a set of fields arranged in a row.

    Arguments:
        wks_record {dict} -- Dictionary with loaded file and parsed record.
        header_fields {List[str]} -- Ordered list of field names.

    Returns:
        Tuple[int, int] -- 1 based index for the found row's index, 1 based column index of the
            first values.
    """
    header_row = None
    column = 1

    # Search for the sample information until found,or you run out of space to look.
    while (
        not header_row
        and column < wks_record["wks"].max_column - len(header_fields) + 2
    ):
        for cell in wks_record["wks"].iter_rows(
            min_col=column, max_col=column, max_row=wks_record["wks"].max_column
        ):
            value = str(cell[0].value) if cell[0].value else ""
            if value.lower() == header_fields[0]:
                header_row = cell[0].row
                return header_row, column
        column += 1
    return header_row, column


def extract_row_sample(generator: Generator) -> List[dict]:
    """
    Returns a row of olink_biorepository_2 samples as a list of dicts with each value paired to
    its header.

    Arguments:
        generator {Generator} -- Generator iterating over the sample rows.

    Returns:
        List[dict] -- List of sample type dictionary objects each representing a row.
    """
    entry_list = []
    sample_headers = [SAMPLE_DESCRIPTION[key] for key in SAMPLE_DESCRIPTION]
    null_row = False
    try:
        while not null_row:
            row = next(generator)
            sample_entry = {}
            for key, cell_value in zip(sample_headers, row):
                sample_entry[key] = cast_cell_value(cell_value)

            # Keep going until a blank row is found.
            if all(sample_entry[key] is None for key in sample_entry):
                null_row = True
            else:
                entry_list.append(sample_entry)
        return entry_list
    except StopIteration:
        return entry_list


def process_clinical_metadata(path: str, context: RecordContext) -> dict:
    """
    Function for dealing with olink metadata xlsx.

    Arguments:
        path {str} -- File path.
        context {RecordContext} -- Background information about the trial,
    Returns:
        dict -- Mongo formatted record.
    """
    metadata_record = {
        "trial": context.trial,
        "assay": context.assay,
        "record_id": context.record,
        "validation_errors": [],
        "samples": [],
    }
    header_fields = [key for key in SAMPLE_DESCRIPTION]
    try:
        xlsx_path = add_file_extension(path, "xlsx")
        # This needs to be adapted to handle multi-sheet workbooks.
        wks_record = {"record": metadata_record, "wks": load_workbook(xlsx_path).active}

        # Handle all information presented in side by side columns of field_name:value type.
        for sub_table in FIELD_NAME_LIST:
            metadata_record.update(parse_matched_column(sub_table, wks_record))

        # Find the sample row.
        header_row, column = find_row_wise(wks_record, header_fields)

        # Check to make sure all the expected headers are there.
        validate_row((header_fields, column, header_row), wks_record)

        # Get cell values.
        metadata_record["samples"] = extract_row_sample(
            wks_record["wks"].iter_rows(
                min_row=header_row + 1,
                min_col=column,
                max_col=column + len(header_fields) - 1,
                max_row=wks_record["wks"].max_column
            )
        )
        lazy_remove(xlsx_path)
        return metadata_record
    except InvalidFileException as err:
        bad_xlsx(metadata_record["validation_errors"], xlsx_path, err)


def process_olink_npx(path: str, context: RecordContext) -> dict:
    """
    Processes an olink npx file and creates a record.

    Arguments:
        path {str} -- Location of the file.
        context {RecordContext} -- Context object containing assay/trial/parent record ID.

    Returns:
        dict -- Olink data formatted for mongodb.
    """
    olink_record = {
        "trial": context.trial,
        "assay": context.assay,
        "record_id": context.record,
        "validation_errors": [],
    }
    xlsx_path = add_file_extension(path, "xlsx")

    try:
        # Load workbook, create iterators to figure out where the relevant information is stored.
        wks = load_workbook(xlsx_path).active
        wks_record = {"record": olink_record, "wks": wks}

        olink_row = validate_column((OLINK_FIRST_COLUMN, 1), wks_record)[1]
        cell_gen = wks.iter_rows(max_col=1, min_row=olink_row + 1)
        # Big function to generate a bunch of important information.
        samples_start_row, samples_end_row, mdf_row, panel_coords, sample_ids = process_sample_rows(
            cell_gen
        )

        if panel_coords:
            olink_record["ol_panel_type"] = wks[panel_coords][1].value
        else:
            olink_record["ol_panel_type"] = None
            olink_record["validation_errors"].append(
                {
                    "explanation": ("Unable to determine panel type"),
                    "affected_paths": ["ol_panel_type"],
                    "raw_or_parse": "PARSE",
                    "severity": "WARNING",
                }
            )

        if not mdf_row:
            olink_record["validation_errors"].append(
                {
                    "explanation": (
                        "The field names in the first column do not match expected"
                        "names"
                    ),
                    "affected_paths": [],
                    "raw_or_parse": "PARSE",
                    "severity": "CRITICAL",
                }
            )

        # Get all of the data needed to create the "assay" objects.
        olink_record["ol_assay"] = extract_assay_data(
            AssayColumn(
                wks.iter_cols(
                    min_col=2,
                    max_col=wks.max_column - 2,
                    min_row=olink_row - 3,
                    max_row=olink_row,
                ),
                wks.iter_cols(
                    min_col=2,
                    max_col=wks.max_column - 2,
                    min_row=mdf_row - 1,
                    max_row=mdf_row,
                ),
                wks.iter_cols(
                    min_col=2,
                    max_col=wks.max_column - 2,
                    min_row=samples_start_row,
                    max_row=samples_end_row,
                ),
            ),
            sample_ids,
            olink_record["validation_errors"],
        )

        # Get the sample-specific information.
        olink_record["samples"] = validate_qc_info(
            SampleInfo(
                wks.iter_cols(
                    min_col=wks.max_column - 1,
                    max_col=wks.max_column,
                    min_row=samples_start_row,
                    max_row=samples_end_row,
                ),
                wks.iter_cols(
                    min_col=wks.max_column,
                    max_col=wks.max_column,
                    min_row=samples_start_row,
                    max_row=samples_end_row,
                ),
                sample_ids,
            ),
            olink_record["validation_errors"],
        )

        # Get NPX Manager Version
        olink_record["npx_m_ver"] = wks["B1"].value

    except InvalidFileException as err:
        bad_xlsx(olink_record["validation_errors"], xlsx_path, err)
    except IndexError:
        log = "Error processing the file"
        logging.error({"message": log, "category": "ERROR-CELERY-FAIR-IMPORT"})
        olink_record["validation_errors"].append(
            mk_error(
                "An index that does not exist was accessed.", [], severity="CRITICAL"
            )
        )
    lazy_remove(xlsx_path)
    return olink_record
